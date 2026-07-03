import io

from openpyxl import Workbook

from app.jira.models import ScopeDataBundle
from app.reports.excel import sheets


class ExcelReportBuilder:
    """Builds the 3-sheet Jira export workbook entirely in memory — never
    touches disk, so the FastAPI endpoint can stream it straight back."""

    def build(self, bundle: ScopeDataBundle) -> io.BytesIO:
        wb = Workbook()

        raw_ws = wb.active
        raw_ws.title = "Sheet 1 - Jira Export"
        sheets.write_raw_sheet(raw_ws, bundle.current_period)

        type_ws = wb.create_sheet("Issue Type - Pivot")
        sheets.write_pivot_by_issue_type(type_ws, bundle.current_period)

        priority_ws = wb.create_sheet("Priority - Pivot")
        sheets.write_pivot_by_priority(priority_ws, bundle.current_period)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
